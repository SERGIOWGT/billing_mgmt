from dataclasses import dataclass
import openpyxl as xl
from src.domain.entities.alojamentos import Alojamento
from src.domain.enums.concessionaria_enum import ConcessionariaEnum

@dataclass
class ExcelReader:

    @staticmethod
    def get_database(file_name):
        wb = xl.load_workbook(file_name)
        sheet_base = wb['Planilha1']
        rows = sheet_base.max_row
        
        alojamentos = []
        for row in range(3, rows):
            nome = sheet_base.cell(row, 3).value
            diretorio = sheet_base.cell(row, 4).value
            for empresa in [x for x in list(ConcessionariaEnum) if x != ConcessionariaEnum.NADA]:
                cliente = sheet_base.cell(row, 2 + (3 * empresa)).value
                conta = sheet_base.cell(row, 3 + (3 * empresa)).value
                local = sheet_base.cell(row, 4 + (3 * empresa)).value

                if cliente or conta or local:
                    alojamentos.append(Alojamento(empresa,nome, diretorio, str(cliente), str(conta), str(local)))

        return alojamentos
